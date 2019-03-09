import mysql.connector
import time
from openpyxl import load_workbook

mydb = mysql.connector.connect(
	host="127.0.0.1",
	user="root",
	passwd="BXY5201314",
	database="gene"
	)

mycursor = mydb.cursor()

colum_names = "GR_ID, SecId, FundId, OldSecID, OldGRID, CUSIP, MFVA, Name, Ticker, InceptDt, MSCat, MSRat3Y, MSRat5Y, MSRat10Y, MSRatAll, AUMShr, hldofInv, OldestSh, Yld12MD, Yld12M, SECYld, SECYldD, Fee12b1, NExpSAR, NExpSARD, NExpPrs, NExpPrsD, BrkAvail, MgrName, Bench, Bench_ID, TR_YTD, TR_3M, TR_6M, TR_9M, TR_1Y, TR_2Y, TR_3Y, TR_4Y, TR_5Y, TR_10Y, TR_15Y, TR_2017, TR_2016, TR_2015, TR_2014, TR_2013, TR_2012, TR_2011, TR_2010, TR_2009, TR_2008, TR_2007, GTR_YTD, GTR_3M, GTR_6M, GTR_9M, GTR_1Y, GTR_2Y, GTR_3Y, GTR_4Y, GTR_5Y, GTR_10Y, GTR_15Y, GTR_2017, GTR_2016, GTR_2015, GTR_2014, GTR_2013, GTR_2012, GTR_2011, GTR_2010, GTR_2009, GTR_2008, GTR_2007, Alpha3, ExRet3, Sharpe3, InfoRat3, Beta3, Stdev3, R23, UpCap3, DnCap3, TrckErr3, FundUp3, FundDn3, BenchUp3, BenchDn3, BatAvg3, OutPNum3"
sqlFormula = "INSERT INTO blog_shares (" + colum_names + ") VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

start_time = time.time()
print("- opening file...")
# ref: read Excel file: https://www.datacamp.com/community/tutorials/python-excel-tutorial
wb = load_workbook("./Share_Data.xlsx")
# print(wb.get_sheet_names())
print("- opened in: [ %s ] seconds" % (time.time() - start_time))
start_time = time.time()

sheet = wb.get_sheet_by_name('Share')
print("-- reading sheet: %s" % sheet.title)

for r in range(12, sheet.max_row + 1): # (2, sheet.max_row + 1), row and col in Excel starts at 1
	GR_ID	= sheet.cell(r,	2).value
	SecId	= sheet.cell(r,	3).value
	FundId	= sheet.cell(r,	4).value
	OldSecID	= sheet.cell(r,	5).value
	OldGRID	= sheet.cell(r,	6).value
	CUSIP	= sheet.cell(r,	7).value
	MFVA	= sheet.cell(r,	8).value
	Name	= sheet.cell(r,	9).value
	Ticker	= sheet.cell(r,	10).value
	InceptDt	= sheet.cell(r,	11).value
	MSCat	= sheet.cell(r,	12).value
	MSRat3Y	= sheet.cell(r,	13).value
	MSRat5Y	= sheet.cell(r,	14).value
	MSRat10Y	= sheet.cell(r,	15).value
	MSRatAll	= sheet.cell(r,	16).value
	AUMShr	= sheet.cell(r,	17).value
	hldofInv	= sheet.cell(r,	18).value
	OldestSh	= sheet.cell(r,	19).value
	Yld12MD	= sheet.cell(r,	20).value
	Yld12M	= sheet.cell(r,	21).value
	SECYld	= sheet.cell(r,	22).value
	SECYldD	= sheet.cell(r,	23).value
	Fee12b1	= sheet.cell(r,	24).value
	NExpSAR	= sheet.cell(r,	25).value
	NExpSARD	= sheet.cell(r,	26).value
	NExpPrs	= sheet.cell(r,	27).value
	NExpPrsD	= sheet.cell(r,	28).value
	BrkAvail	= sheet.cell(r,	29).value
	MgrName	= sheet.cell(r,	30).value
	Bench	= sheet.cell(r,	31).value
	Bench_ID	= sheet.cell(r,	32).value
	TR_YTD	= sheet.cell(r,	33).value
	TR_3M	= sheet.cell(r,	34).value
	TR_6M	= sheet.cell(r,	35).value
	TR_9M	= sheet.cell(r,	36).value
	TR_1Y	= sheet.cell(r,	37).value
	TR_2Y	= sheet.cell(r,	38).value
	TR_3Y	= sheet.cell(r,	39).value
	TR_4Y	= sheet.cell(r,	40).value
	TR_5Y	= sheet.cell(r,	41).value
	TR_10Y	= sheet.cell(r,	42).value
	TR_15Y	= sheet.cell(r,	43).value
	TR_2017	= sheet.cell(r,	44).value
	TR_2016	= sheet.cell(r,	45).value
	TR_2015	= sheet.cell(r,	46).value
	TR_2014	= sheet.cell(r,	47).value
	TR_2013	= sheet.cell(r,	48).value
	TR_2012	= sheet.cell(r,	49).value
	TR_2011	= sheet.cell(r,	50).value
	TR_2010	= sheet.cell(r,	51).value
	TR_2009	= sheet.cell(r,	52).value
	TR_2008	= sheet.cell(r,	53).value
	TR_2007	= sheet.cell(r,	54).value
	GTR_YTD	= sheet.cell(r,	55).value
	GTR_3M	= sheet.cell(r,	56).value
	GTR_6M	= sheet.cell(r,	57).value
	GTR_9M	= sheet.cell(r,	58).value
	GTR_1Y	= sheet.cell(r,	59).value
	GTR_2Y = None if len(str(sheet.cell(r,	60).value)) < 2 else sheet.cell(r,	60).value # Incorrect decimal value: '' for column,
	GTR_3Y	= None if len(str(sheet.cell(r,	61).value)) < 2 else sheet.cell(r,	61).value # and excel unable to find ' in cell, so...
	GTR_4Y	= None if len(str(sheet.cell(r,	62).value)) < 2 else sheet.cell(r,	62).value
	GTR_5Y	= None if len(str(sheet.cell(r,	63).value)) < 2 else sheet.cell(r,	63).value
	GTR_10Y	= sheet.cell(r,	64).value
	GTR_15Y	= sheet.cell(r,	65).value
	GTR_2017	= sheet.cell(r,	66).value
	GTR_2016	= sheet.cell(r,	67).value
	GTR_2015	= sheet.cell(r,	68).value
	GTR_2014	= sheet.cell(r,	69).value
	GTR_2013	= sheet.cell(r,	70).value
	GTR_2012	= sheet.cell(r,	71).value
	GTR_2011	= sheet.cell(r,	72).value
	GTR_2010	= sheet.cell(r,	73).value
	GTR_2009	= sheet.cell(r,	74).value
	GTR_2008	= sheet.cell(r,	75).value
	GTR_2007	= sheet.cell(r,	76).value
	Alpha3	= sheet.cell(r,	77).value
	ExRet3	= sheet.cell(r,	78).value
	Sharpe3	= sheet.cell(r,	79).value
	InfoRat3	= sheet.cell(r,	80).value
	Beta3	= sheet.cell(r,	81).value
	Stdev3	= sheet.cell(r,	82).value
	R23	= sheet.cell(r,	83).value
	UpCap3	= sheet.cell(r,	84).value
	DnCap3	= sheet.cell(r,	85).value
	TrckErr3	= sheet.cell(r,	86).value
	FundUp3	= sheet.cell(r,	87).value
	FundDn3	= sheet.cell(r,	88).value
	BenchUp3	= sheet.cell(r,	89).value
	BenchDn3	= sheet.cell(r,	90).value
	BatAvg3	= sheet.cell(r,	91).value
	OutPNum3	= sheet.cell(r,	92).value

	values = (GR_ID, SecId, FundId, OldSecID, OldGRID, CUSIP, MFVA, Name, Ticker, InceptDt, MSCat, MSRat3Y, MSRat5Y, MSRat10Y, MSRatAll, AUMShr, hldofInv, OldestSh, Yld12MD, Yld12M, SECYld, SECYldD, Fee12b1, NExpSAR, NExpSARD, NExpPrs, NExpPrsD, BrkAvail, MgrName, Bench, Bench_ID, TR_YTD, TR_3M, TR_6M, TR_9M, TR_1Y, TR_2Y, TR_3Y, TR_4Y, TR_5Y, TR_10Y, TR_15Y, TR_2017, TR_2016, TR_2015, TR_2014, TR_2013, TR_2012, TR_2011, TR_2010, TR_2009, TR_2008, TR_2007, GTR_YTD, GTR_3M, GTR_6M, GTR_9M, GTR_1Y, GTR_2Y, GTR_3Y, GTR_4Y, GTR_5Y, GTR_10Y, GTR_15Y, GTR_2017, GTR_2016, GTR_2015, GTR_2014, GTR_2013, GTR_2012, GTR_2011, GTR_2010, GTR_2009, GTR_2008, GTR_2007, Alpha3, ExRet3, Sharpe3, InfoRat3, Beta3, Stdev3, R23, UpCap3, DnCap3, TrckErr3, FundUp3, FundDn3, BenchUp3, BenchDn3, BatAvg3, OutPNum3)

	# print(r)
	# print("--- inserting val = %s", values)
	mycursor.execute(sqlFormula, values)

print("--- finished inserting in: [ %s ] seconds" % (time.time() - start_time))
print("--- close db cursor.")
mycursor.close()

# print("--- will commit to db...")
mydb.commit()
mydb.close()
print("--- close db.")



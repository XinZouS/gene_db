import mysql.connector
import time
from openpyxl import load_workbook
import db_config


mydb = mysql.connector.connect(
	host=db_config.local_host,
	user=db_config.user,
	passwd=db_config.passwd,
	database=db_config.database	
	)

mycursor = mydb.cursor()

colum_names = "FundId, Fund, SubAdvisor, SubAdvisorParent, AdvisorParent, SubAdvised, AgrmStart, AgrmEnd, SubStart, SubEnd, SubAlloc, SubAUM, FundAUM, EffSub, SubSched3"
sqlFormula = "INSERT INTO research_SubAdv (" + colum_names + ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

start_time = time.time()
print("- opening file...")
# ref: read Excel file: https://www.datacamp.com/community/tutorials/python-excel-tutorial
wb = load_workbook("./SubAdv_Sched.xlsx")
# print(wb.get_sheet_names())
print("- opened in: [ %s ] seconds" % (time.time() - start_time))
start_time = time.time()

sheet = wb.get_sheet_by_name('SubAdv_Sched')
print("-- reading sheet: %s" % sheet.title)


def roundFloat(val, decimals):
	if val:
		return round(float(val), decimals)
	return val


def toFloat_3(val):
	if val is not None:
		return '%.3f' % val
	else:
		return None


def toFloat_6(val):
	if val is not None:
		return '%.6f' % val
	else:
		return None


i = 1
total = sheet.max_row + 1 - 12

for r in range(2, sheet.max_row + 1): # (2, sheet.max_row + 1), row and col in Excel starts at 1
	FundId	= sheet.cell(r,	1).value
	Fund	= sheet.cell(r,	2).value
	SubAdvisor	= sheet.cell(r,	3).value
	SubAdvisorParent	= sheet.cell(r,	4).value
	AdvisorParent	= sheet.cell(r,	5).value
	SubAdvised	= sheet.cell(r,	6).value
	AgrmStart	= sheet.cell(r,	7).value
	AgrmEnd	= sheet.cell(r,	8).value
	SubStart	= sheet.cell(r,	9).value
	SubEnd	= sheet.cell(r,	10).value
	SubAlloc= toFloat_3(sheet.cell(r,11).value)
	SubAUM	= toFloat_6(sheet.cell(r,12).value)
	FundAUM	= toFloat_6(sheet.cell(r,13).value)
	EffSub	= toFloat_3(sheet.cell(r,14).value)
	SubSched3	= sheet.cell(r,	15).value

	values = (FundId, Fund, SubAdvisor, SubAdvisorParent, AdvisorParent, SubAdvised, AgrmStart, AgrmEnd, SubStart, SubEnd, SubAlloc, SubAUM, FundAUM, EffSub, SubSched3)

	# print(r)
	# print("--- inserting val = ", values)
	print('%s / %s' % (i,total))
	i += 1
	if Fund:
		mycursor.execute(sqlFormula, values)


print("--- finished inserting in: [ %s ] seconds" % (time.time() - start_time))
print("--- close db cursor.")
mycursor.close()

print("--- will commit to db...")
mydb.commit()

print("--- close db.")
mydb.close()



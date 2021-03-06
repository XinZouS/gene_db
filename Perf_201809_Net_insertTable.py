import mysql.connector
import time
from openpyxl import load_workbook
import db_config


mydb = mysql.connector.connect(
	host=db_config.local_host,
	user=db_config.user,
	passwd=db_config.passwd,
	database=db_config.database	
	)

mycursor = mydb.cursor()

colum_names = "Name, SecId, Category, TR201809, TR201808, TR201807, TR201806, TR201805, TR201804, TR201803, TR201802, TR201801, TR201712, TR201711, TR201710, TR201709, TR201708, TR201707, TR201706, TR201705, TR201704, TR201703, TR201702, TR201701, TR201612, TR201611, TR201610, TR201609, TR201608, TR201607, TR201606, TR201605, TR201604, TR201603, TR201602, TR201601, TR201512, TR201511, TR201510, TR201509, TR201508, TR201507, TR201506, TR201505, TR201504, TR201503, TR201502, TR201501, TR201412, TR201411, TR201410, TR201409, TR201408, TR201407, TR201406, TR201405, TR201404, TR201403, TR201402, TR201401, TR201312, TR201311, TR201310, TR201309, TR201308, TR201307, TR201306, TR201305, TR201304, TR201303, TR201302, TR201301, TR201212, TR201211, TR201210, TR201209, TR201208, TR201207, TR201206, TR201205, TR201204, TR201203, TR201202, TR201201, TR201112, TR201111, TR201110, TR201109, TR201108, TR201107, TR201106, TR201105, TR201104, TR201103, TR201102, TR201101, TR201012, TR201011, TR201010, TR201009, TR201008, TR201007, TR201006, TR201005, TR201004, TR201003, TR201002, TR201001, TR200912, TR200911, TR200910, TR200909, TR200908, TR200907, TR200906, TR200905, TR200904, TR200903, TR200902, TR200901, TR200812, TR200811, TR200810, TR200809, TR200808, TR200807, TR200806, TR200805, TR200804, TR200803, TR200802, TR200801, TR200712, TR200711, TR200710, TR200709, TR200708, TR200707, TR200706, TR200705, TR200704, TR200703, TR200702, TR200701, TR200612, TR200611, TR200610, TR200609, TR200608, TR200607, TR200606, TR200605, TR200604, TR200603, TR200602, TR200601, TR200512, TR200511, TR200510, TR200509, TR200508, TR200507, TR200506, TR200505, TR200504, TR200503, TR200502, TR200501, TR200412, TR200411, TR200410, TR200409, TR200408, TR200407, TR200406, TR200405, TR200404, TR200403, TR200402, TR200401, TR200312, TR200311, TR200310, TR200309, TR200308, TR200307, TR200306, TR200305, TR200304, TR200303, TR200302, TR200301, TR200212, TR200211, TR200210, TR200209, TR200208, TR200207, TR200206, TR200205, TR200204, TR200203, TR200202, TR200201, TR200112, TR200111, TR200110, TR200109, TR200108, TR200107, TR200106, TR200105, TR200104, TR200103, TR200102, TR200101, TR200012, TR200011, TR200010, TR200009, TR200008, TR200007, TR200006, TR200005, TR200004, TR200003, TR200002, TR200001, TR199912, TR199911, TR199910, TR199909, TR199908, TR199907, TR199906, TR199905, TR199904, TR199903, TR199902, TR199901, TR199812, TR199811, TR199810, TR199809, TR199808, TR199807, TR199806, TR199805, TR199804, TR199803, TR199802, TR199801, TR199712, TR199711, TR199710, TR199709, TR199708, TR199707, TR199706, TR199705, TR199704, TR199703, TR199702, TR199701, TR199612, TR199611, TR199610, TR199609, TR199608, TR199607, TR199606, TR199605, TR199604, TR199603, TR199602, TR199601, TR199512, TR199511, TR199510, TR199509, TR199508, TR199507, TR199506, TR199505, TR199504, TR199503, TR199502, TR199501, TR199412, TR199411, TR199410, TR199409, TR199408, TR199407, TR199406"
sqlFormula = "INSERT INTO perf_net (" + colum_names + ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

start_time = time.time()
print("- opening file...")
# ref: read Excel file: https://www.datacamp.com/community/tutorials/python-excel-tutorial
wb = load_workbook("./Perf_201809_Net.xlsx")
# print(wb.get_sheet_names())
print("- opened in: [ %s ] seconds" % (time.time() - start_time))
start_time = time.time()

sheet = wb.get_sheet_by_name('Net')
print("-- reading sheet: %s" % sheet.title)

for r in range(2, sheet.max_row + 1): # (2, sheet.max_row + 1), row and col in Excel starts at 1
	Name = sheet.cell(r, 1 ).value
	SecId = sheet.cell(r, 2 ).value
	Category = sheet.cell(r, 3 ).value
	TR201809 = sheet.cell(r, 4 ).value
	TR201808 = sheet.cell(r, 5 ).value
	TR201807 = sheet.cell(r, 6 ).value
	TR201806 = sheet.cell(r, 7 ).value
	TR201805 = sheet.cell(r, 8 ).value
	TR201804 = sheet.cell(r, 9 ).value
	TR201803 = sheet.cell(r, 10 ).value
	TR201802 = sheet.cell(r, 11 ).value
	TR201801 = sheet.cell(r, 12 ).value
	TR201712 = sheet.cell(r, 13 ).value
	TR201711 = sheet.cell(r, 14 ).value
	TR201710 = sheet.cell(r, 15 ).value
	TR201709 = sheet.cell(r, 16 ).value
	TR201708 = sheet.cell(r, 17 ).value
	TR201707 = sheet.cell(r, 18 ).value
	TR201706 = sheet.cell(r, 19 ).value
	TR201705 = sheet.cell(r, 20 ).value
	TR201704 = sheet.cell(r, 21 ).value
	TR201703 = sheet.cell(r, 22 ).value
	TR201702 = sheet.cell(r, 23 ).value
	TR201701 = sheet.cell(r, 24 ).value
	TR201612 = sheet.cell(r, 25 ).value
	TR201611 = sheet.cell(r, 26 ).value
	TR201610 = sheet.cell(r, 27 ).value
	TR201609 = sheet.cell(r, 28 ).value
	TR201608 = sheet.cell(r, 29 ).value
	TR201607 = sheet.cell(r, 30 ).value
	TR201606 = sheet.cell(r, 31 ).value
	TR201605 = sheet.cell(r, 32 ).value
	TR201604 = sheet.cell(r, 33 ).value
	TR201603 = sheet.cell(r, 34 ).value
	TR201602 = sheet.cell(r, 35 ).value
	TR201601 = sheet.cell(r, 36 ).value
	TR201512 = sheet.cell(r, 37 ).value
	TR201511 = sheet.cell(r, 38 ).value
	TR201510 = sheet.cell(r, 39 ).value
	TR201509 = sheet.cell(r, 40 ).value
	TR201508 = sheet.cell(r, 41 ).value
	TR201507 = sheet.cell(r, 42 ).value
	TR201506 = sheet.cell(r, 43 ).value
	TR201505 = sheet.cell(r, 44 ).value
	TR201504 = sheet.cell(r, 45 ).value
	TR201503 = sheet.cell(r, 46 ).value
	TR201502 = sheet.cell(r, 47 ).value
	TR201501 = sheet.cell(r, 48 ).value
	TR201412 = sheet.cell(r, 49 ).value
	TR201411 = sheet.cell(r, 50 ).value
	TR201410 = sheet.cell(r, 51 ).value
	TR201409 = sheet.cell(r, 52 ).value
	TR201408 = sheet.cell(r, 53 ).value
	TR201407 = sheet.cell(r, 54 ).value
	TR201406 = sheet.cell(r, 55 ).value
	TR201405 = sheet.cell(r, 56 ).value
	TR201404 = sheet.cell(r, 57 ).value
	TR201403 = sheet.cell(r, 58 ).value
	TR201402 = sheet.cell(r, 59 ).value
	TR201401 = sheet.cell(r, 60 ).value
	TR201312 = sheet.cell(r, 61 ).value
	TR201311 = sheet.cell(r, 62 ).value
	TR201310 = sheet.cell(r, 63 ).value
	TR201309 = sheet.cell(r, 64 ).value
	TR201308 = sheet.cell(r, 65 ).value
	TR201307 = sheet.cell(r, 66 ).value
	TR201306 = sheet.cell(r, 67 ).value
	TR201305 = sheet.cell(r, 68 ).value
	TR201304 = sheet.cell(r, 69 ).value
	TR201303 = sheet.cell(r, 70 ).value
	TR201302 = sheet.cell(r, 71 ).value
	TR201301 = sheet.cell(r, 72 ).value
	TR201212 = sheet.cell(r, 73 ).value
	TR201211 = sheet.cell(r, 74 ).value
	TR201210 = sheet.cell(r, 75 ).value
	TR201209 = sheet.cell(r, 76 ).value
	TR201208 = sheet.cell(r, 77 ).value
	TR201207 = sheet.cell(r, 78 ).value
	TR201206 = sheet.cell(r, 79 ).value
	TR201205 = sheet.cell(r, 80 ).value
	TR201204 = sheet.cell(r, 81 ).value
	TR201203 = sheet.cell(r, 82 ).value
	TR201202 = sheet.cell(r, 83 ).value
	TR201201 = sheet.cell(r, 84 ).value
	TR201112 = sheet.cell(r, 85 ).value
	TR201111 = sheet.cell(r, 86 ).value
	TR201110 = sheet.cell(r, 87 ).value
	TR201109 = sheet.cell(r, 88 ).value
	TR201108 = sheet.cell(r, 89 ).value
	TR201107 = sheet.cell(r, 90 ).value
	TR201106 = sheet.cell(r, 91 ).value
	TR201105 = sheet.cell(r, 92 ).value
	TR201104 = sheet.cell(r, 93 ).value
	TR201103 = sheet.cell(r, 94 ).value
	TR201102 = sheet.cell(r, 95 ).value
	TR201101 = sheet.cell(r, 96 ).value
	TR201012 = sheet.cell(r, 97 ).value
	TR201011 = sheet.cell(r, 98 ).value
	TR201010 = sheet.cell(r, 99 ).value
	TR201009 = sheet.cell(r, 100 ).value
	TR201008 = sheet.cell(r, 101 ).value
	TR201007 = sheet.cell(r, 102 ).value
	TR201006 = sheet.cell(r, 103 ).value
	TR201005 = sheet.cell(r, 104 ).value
	TR201004 = sheet.cell(r, 105 ).value
	TR201003 = sheet.cell(r, 106 ).value
	TR201002 = sheet.cell(r, 107 ).value
	TR201001 = sheet.cell(r, 108 ).value
	TR200912 = sheet.cell(r, 109 ).value
	TR200911 = sheet.cell(r, 110 ).value
	TR200910 = sheet.cell(r, 111 ).value
	TR200909 = sheet.cell(r, 112 ).value
	TR200908 = sheet.cell(r, 113 ).value
	TR200907 = sheet.cell(r, 114 ).value
	TR200906 = sheet.cell(r, 115 ).value
	TR200905 = sheet.cell(r, 116 ).value
	TR200904 = sheet.cell(r, 117 ).value
	TR200903 = sheet.cell(r, 118 ).value
	TR200902 = sheet.cell(r, 119 ).value
	TR200901 = sheet.cell(r, 120 ).value
	TR200812 = sheet.cell(r, 121 ).value
	TR200811 = sheet.cell(r, 122 ).value
	TR200810 = sheet.cell(r, 123 ).value
	TR200809 = sheet.cell(r, 124 ).value
	TR200808 = sheet.cell(r, 125 ).value
	TR200807 = sheet.cell(r, 126 ).value
	TR200806 = sheet.cell(r, 127 ).value
	TR200805 = sheet.cell(r, 128 ).value
	TR200804 = sheet.cell(r, 129 ).value
	TR200803 = sheet.cell(r, 130 ).value
	TR200802 = sheet.cell(r, 131 ).value
	TR200801 = sheet.cell(r, 132 ).value
	TR200712 = sheet.cell(r, 133 ).value
	TR200711 = sheet.cell(r, 134 ).value
	TR200710 = sheet.cell(r, 135 ).value
	TR200709 = sheet.cell(r, 136 ).value
	TR200708 = sheet.cell(r, 137 ).value
	TR200707 = sheet.cell(r, 138 ).value
	TR200706 = sheet.cell(r, 139 ).value
	TR200705 = sheet.cell(r, 140 ).value
	TR200704 = sheet.cell(r, 141 ).value
	TR200703 = sheet.cell(r, 142 ).value
	TR200702 = sheet.cell(r, 143 ).value
	TR200701 = sheet.cell(r, 144 ).value
	TR200612 = sheet.cell(r, 145 ).value
	TR200611 = sheet.cell(r, 146 ).value
	TR200610 = sheet.cell(r, 147 ).value
	TR200609 = sheet.cell(r, 148 ).value
	TR200608 = sheet.cell(r, 149 ).value
	TR200607 = sheet.cell(r, 150 ).value
	TR200606 = sheet.cell(r, 151 ).value
	TR200605 = sheet.cell(r, 152 ).value
	TR200604 = sheet.cell(r, 153 ).value
	TR200603 = sheet.cell(r, 154 ).value
	TR200602 = sheet.cell(r, 155 ).value
	TR200601 = sheet.cell(r, 156 ).value
	TR200512 = sheet.cell(r, 157 ).value
	TR200511 = sheet.cell(r, 158 ).value
	TR200510 = sheet.cell(r, 159 ).value
	TR200509 = sheet.cell(r, 160 ).value
	TR200508 = sheet.cell(r, 161 ).value
	TR200507 = sheet.cell(r, 162 ).value
	TR200506 = sheet.cell(r, 163 ).value
	TR200505 = sheet.cell(r, 164 ).value
	TR200504 = sheet.cell(r, 165 ).value
	TR200503 = sheet.cell(r, 166 ).value
	TR200502 = sheet.cell(r, 167 ).value
	TR200501 = sheet.cell(r, 168 ).value
	TR200412 = sheet.cell(r, 169 ).value
	TR200411 = sheet.cell(r, 170 ).value
	TR200410 = sheet.cell(r, 171 ).value
	TR200409 = sheet.cell(r, 172 ).value
	TR200408 = sheet.cell(r, 173 ).value
	TR200407 = sheet.cell(r, 174 ).value
	TR200406 = sheet.cell(r, 175 ).value
	TR200405 = sheet.cell(r, 176 ).value
	TR200404 = sheet.cell(r, 177 ).value
	TR200403 = sheet.cell(r, 178 ).value
	TR200402 = sheet.cell(r, 179 ).value
	TR200401 = sheet.cell(r, 180 ).value
	TR200312 = sheet.cell(r, 181 ).value
	TR200311 = sheet.cell(r, 182 ).value
	TR200310 = sheet.cell(r, 183 ).value
	TR200309 = sheet.cell(r, 184 ).value
	TR200308 = sheet.cell(r, 185 ).value
	TR200307 = sheet.cell(r, 186 ).value
	TR200306 = sheet.cell(r, 187 ).value
	TR200305 = sheet.cell(r, 188 ).value
	TR200304 = sheet.cell(r, 189 ).value
	TR200303 = sheet.cell(r, 190 ).value
	TR200302 = sheet.cell(r, 191 ).value
	TR200301 = sheet.cell(r, 192 ).value
	TR200212 = sheet.cell(r, 193 ).value
	TR200211 = sheet.cell(r, 194 ).value
	TR200210 = sheet.cell(r, 195 ).value
	TR200209 = sheet.cell(r, 196 ).value
	TR200208 = sheet.cell(r, 197 ).value
	TR200207 = sheet.cell(r, 198 ).value
	TR200206 = sheet.cell(r, 199 ).value
	TR200205 = sheet.cell(r, 200 ).value
	TR200204 = sheet.cell(r, 201 ).value
	TR200203 = sheet.cell(r, 202 ).value
	TR200202 = sheet.cell(r, 203 ).value
	TR200201 = sheet.cell(r, 204 ).value
	TR200112 = sheet.cell(r, 205 ).value
	TR200111 = sheet.cell(r, 206 ).value
	TR200110 = sheet.cell(r, 207 ).value
	TR200109 = sheet.cell(r, 208 ).value
	TR200108 = sheet.cell(r, 209 ).value
	TR200107 = sheet.cell(r, 210 ).value
	TR200106 = sheet.cell(r, 211 ).value
	TR200105 = sheet.cell(r, 212 ).value
	TR200104 = sheet.cell(r, 213 ).value
	TR200103 = sheet.cell(r, 214 ).value
	TR200102 = sheet.cell(r, 215 ).value
	TR200101 = sheet.cell(r, 216 ).value
	TR200012 = sheet.cell(r, 217 ).value
	TR200011 = sheet.cell(r, 218 ).value
	TR200010 = sheet.cell(r, 219 ).value
	TR200009 = sheet.cell(r, 220 ).value
	TR200008 = sheet.cell(r, 221 ).value
	TR200007 = sheet.cell(r, 222 ).value
	TR200006 = sheet.cell(r, 223 ).value
	TR200005 = sheet.cell(r, 224 ).value
	TR200004 = sheet.cell(r, 225 ).value
	TR200003 = sheet.cell(r, 226 ).value
	TR200002 = sheet.cell(r, 227 ).value
	TR200001 = sheet.cell(r, 228 ).value
	TR199912 = sheet.cell(r, 229 ).value
	TR199911 = sheet.cell(r, 230 ).value
	TR199910 = sheet.cell(r, 231 ).value
	TR199909 = sheet.cell(r, 232 ).value
	TR199908 = sheet.cell(r, 233 ).value
	TR199907 = sheet.cell(r, 234 ).value
	TR199906 = sheet.cell(r, 235 ).value
	TR199905 = sheet.cell(r, 236 ).value
	TR199904 = sheet.cell(r, 237 ).value
	TR199903 = sheet.cell(r, 238 ).value
	TR199902 = sheet.cell(r, 239 ).value
	TR199901 = sheet.cell(r, 240 ).value
	TR199812 = sheet.cell(r, 241 ).value
	TR199811 = sheet.cell(r, 242 ).value
	TR199810 = sheet.cell(r, 243 ).value
	TR199809 = sheet.cell(r, 244 ).value
	TR199808 = sheet.cell(r, 245 ).value
	TR199807 = sheet.cell(r, 246 ).value
	TR199806 = sheet.cell(r, 247 ).value
	TR199805 = sheet.cell(r, 248 ).value
	TR199804 = sheet.cell(r, 249 ).value
	TR199803 = sheet.cell(r, 250 ).value
	TR199802 = sheet.cell(r, 251 ).value
	TR199801 = sheet.cell(r, 252 ).value
	TR199712 = sheet.cell(r, 253 ).value
	TR199711 = sheet.cell(r, 254 ).value
	TR199710 = sheet.cell(r, 255 ).value
	TR199709 = sheet.cell(r, 256 ).value
	TR199708 = sheet.cell(r, 257 ).value
	TR199707 = sheet.cell(r, 258 ).value
	TR199706 = sheet.cell(r, 259 ).value
	TR199705 = sheet.cell(r, 260 ).value
	TR199704 = sheet.cell(r, 261 ).value
	TR199703 = sheet.cell(r, 262 ).value
	TR199702 = sheet.cell(r, 263 ).value
	TR199701 = sheet.cell(r, 264 ).value
	TR199612 = sheet.cell(r, 265 ).value
	TR199611 = sheet.cell(r, 266 ).value
	TR199610 = sheet.cell(r, 267 ).value
	TR199609 = sheet.cell(r, 268 ).value
	TR199608 = sheet.cell(r, 269 ).value
	TR199607 = sheet.cell(r, 270 ).value
	TR199606 = sheet.cell(r, 271 ).value
	TR199605 = sheet.cell(r, 272 ).value
	TR199604 = sheet.cell(r, 273 ).value
	TR199603 = sheet.cell(r, 274 ).value
	TR199602 = sheet.cell(r, 275 ).value
	TR199601 = sheet.cell(r, 276 ).value
	TR199512 = sheet.cell(r, 277 ).value
	TR199511 = sheet.cell(r, 278 ).value
	TR199510 = sheet.cell(r, 279 ).value
	TR199509 = sheet.cell(r, 280 ).value
	TR199508 = sheet.cell(r, 281 ).value
	TR199507 = sheet.cell(r, 282 ).value
	TR199506 = sheet.cell(r, 283 ).value
	TR199505 = sheet.cell(r, 284 ).value
	TR199504 = sheet.cell(r, 285 ).value
	TR199503 = sheet.cell(r, 286 ).value
	TR199502 = sheet.cell(r, 287 ).value
	TR199501 = sheet.cell(r, 288 ).value
	TR199412 = sheet.cell(r, 289 ).value
	TR199411 = sheet.cell(r, 290 ).value
	TR199410 = sheet.cell(r, 291 ).value
	TR199409 = sheet.cell(r, 292 ).value
	TR199408 = sheet.cell(r, 293 ).value
	TR199407 = sheet.cell(r, 294 ).value
	TR199406 = sheet.cell(r, 295 ).value
		
	values = (Name, SecId, Category, TR201809, TR201808, TR201807, TR201806, TR201805, TR201804, TR201803, TR201802, TR201801, TR201712, TR201711, TR201710, TR201709, TR201708, TR201707, TR201706, TR201705, TR201704, TR201703, TR201702, TR201701, TR201612, TR201611, TR201610, TR201609, TR201608, TR201607, TR201606, TR201605, TR201604, TR201603, TR201602, TR201601, TR201512, TR201511, TR201510, TR201509, TR201508, TR201507, TR201506, TR201505, TR201504, TR201503, TR201502, TR201501, TR201412, TR201411, TR201410, TR201409, TR201408, TR201407, TR201406, TR201405, TR201404, TR201403, TR201402, TR201401, TR201312, TR201311, TR201310, TR201309, TR201308, TR201307, TR201306, TR201305, TR201304, TR201303, TR201302, TR201301, TR201212, TR201211, TR201210, TR201209, TR201208, TR201207, TR201206, TR201205, TR201204, TR201203, TR201202, TR201201, TR201112, TR201111, TR201110, TR201109, TR201108, TR201107, TR201106, TR201105, TR201104, TR201103, TR201102, TR201101, TR201012, TR201011, TR201010, TR201009, TR201008, TR201007, TR201006, TR201005, TR201004, TR201003, TR201002, TR201001, TR200912, TR200911, TR200910, TR200909, TR200908, TR200907, TR200906, TR200905, TR200904, TR200903, TR200902, TR200901, TR200812, TR200811, TR200810, TR200809, TR200808, TR200807, TR200806, TR200805, TR200804, TR200803, TR200802, TR200801, TR200712, TR200711, TR200710, TR200709, TR200708, TR200707, TR200706, TR200705, TR200704, TR200703, TR200702, TR200701, TR200612, TR200611, TR200610, TR200609, TR200608, TR200607, TR200606, TR200605, TR200604, TR200603, TR200602, TR200601, TR200512, TR200511, TR200510, TR200509, TR200508, TR200507, TR200506, TR200505, TR200504, TR200503, TR200502, TR200501, TR200412, TR200411, TR200410, TR200409, TR200408, TR200407, TR200406, TR200405, TR200404, TR200403, TR200402, TR200401, TR200312, TR200311, TR200310, TR200309, TR200308, TR200307, TR200306, TR200305, TR200304, TR200303, TR200302, TR200301, TR200212, TR200211, TR200210, TR200209, TR200208, TR200207, TR200206, TR200205, TR200204, TR200203, TR200202, TR200201, TR200112, TR200111, TR200110, TR200109, TR200108, TR200107, TR200106, TR200105, TR200104, TR200103, TR200102, TR200101, TR200012, TR200011, TR200010, TR200009, TR200008, TR200007, TR200006, TR200005, TR200004, TR200003, TR200002, TR200001, TR199912, TR199911, TR199910, TR199909, TR199908, TR199907, TR199906, TR199905, TR199904, TR199903, TR199902, TR199901, TR199812, TR199811, TR199810, TR199809, TR199808, TR199807, TR199806, TR199805, TR199804, TR199803, TR199802, TR199801, TR199712, TR199711, TR199710, TR199709, TR199708, TR199707, TR199706, TR199705, TR199704, TR199703, TR199702, TR199701, TR199612, TR199611, TR199610, TR199609, TR199608, TR199607, TR199606, TR199605, TR199604, TR199603, TR199602, TR199601, TR199512, TR199511, TR199510, TR199509, TR199508, TR199507, TR199506, TR199505, TR199504, TR199503, TR199502, TR199501, TR199412, TR199411, TR199410, TR199409, TR199408, TR199407, TR199406)

	# print(r)
	# print("--- inserting val = %s", values)
	mycursor.execute(sqlFormula, values)

print("--- finished inserting in: [ %s ] seconds" % (time.time() - start_time))
print("--- close db cursor.")
mycursor.close()

print("--- will commit to db...")
mydb.commit()

print("--- close db.")
mydb.close()



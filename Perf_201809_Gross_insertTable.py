import mysql.connector
import time
from openpyxl import load_workbook
import db_config


mydb = mysql.connector.connect(
	host=db_config.local_host,
	user=db_config.user,
	passwd=db_config.passwd,
	database=db_config.database	
	)


mycursor = mydb.cursor()

colum_names = "Name, SecId, Category, GTR201809, GTR201808, GTR201807, GTR201806, GTR201805, GTR201804, GTR201803, GTR201802, GTR201801, GTR201712, GTR201711, GTR201710, GTR201709, GTR201708, GTR201707, GTR201706, GTR201705, GTR201704, GTR201703, GTR201702, GTR201701, GTR201612, GTR201611, GTR201610, GTR201609, GTR201608, GTR201607, GTR201606, GTR201605, GTR201604, GTR201603, GTR201602, GTR201601, GTR201512, GTR201511, GTR201510, GTR201509, GTR201508, GTR201507, GTR201506, GTR201505, GTR201504, GTR201503, GTR201502, GTR201501, GTR201412, GTR201411, GTR201410, GTR201409, GTR201408, GTR201407, GTR201406, GTR201405, GTR201404, GTR201403, GTR201402, GTR201401, GTR201312, GTR201311, GTR201310, GTR201309, GTR201308, GTR201307, GTR201306, GTR201305, GTR201304, GTR201303, GTR201302, GTR201301, GTR201212, GTR201211, GTR201210, GTR201209, GTR201208, GTR201207, GTR201206, GTR201205, GTR201204, GTR201203, GTR201202, GTR201201, GTR201112, GTR201111, GTR201110, GTR201109, GTR201108, GTR201107, GTR201106, GTR201105, GTR201104, GTR201103, GTR201102, GTR201101, GTR201012, GTR201011, GTR201010, GTR201009, GTR201008, GTR201007, GTR201006, GTR201005, GTR201004, GTR201003, GTR201002, GTR201001, GTR200912, GTR200911, GTR200910, GTR200909, GTR200908, GTR200907, GTR200906, GTR200905, GTR200904, GTR200903, GTR200902, GTR200901, GTR200812, GTR200811, GTR200810, GTR200809, GTR200808, GTR200807, GTR200806, GTR200805, GTR200804, GTR200803, GTR200802, GTR200801, GTR200712, GTR200711, GTR200710, GTR200709, GTR200708, GTR200707, GTR200706, GTR200705, GTR200704, GTR200703, GTR200702, GTR200701, GTR200612, GTR200611, GTR200610, GTR200609, GTR200608, GTR200607, GTR200606, GTR200605, GTR200604, GTR200603, GTR200602, GTR200601, GTR200512, GTR200511, GTR200510, GTR200509, GTR200508, GTR200507, GTR200506, GTR200505, GTR200504, GTR200503, GTR200502, GTR200501, GTR200412, GTR200411, GTR200410, GTR200409, GTR200408, GTR200407, GTR200406, GTR200405, GTR200404, GTR200403, GTR200402, GTR200401, GTR200312, GTR200311, GTR200310, GTR200309, GTR200308, GTR200307, GTR200306, GTR200305, GTR200304, GTR200303, GTR200302, GTR200301, GTR200212, GTR200211, GTR200210, GTR200209, GTR200208, GTR200207, GTR200206, GTR200205, GTR200204, GTR200203, GTR200202, GTR200201, GTR200112, GTR200111, GTR200110, GTR200109, GTR200108, GTR200107, GTR200106, GTR200105, GTR200104, GTR200103, GTR200102, GTR200101, GTR200012, GTR200011, GTR200010, GTR200009, GTR200008, GTR200007, GTR200006, GTR200005, GTR200004, GTR200003, GTR200002, GTR200001, GTR199912, GTR199911, GTR199910, GTR199909, GTR199908, GTR199907, GTR199906, GTR199905, GTR199904, GTR199903, GTR199902, GTR199901, GTR199812, GTR199811, GTR199810, GTR199809, GTR199808, GTR199807, GTR199806, GTR199805, GTR199804, GTR199803, GTR199802, GTR199801, GTR199712, GTR199711, GTR199710, GTR199709, GTR199708, GTR199707, GTR199706, GTR199705, GTR199704, GTR199703, GTR199702, GTR199701, GTR199612, GTR199611, GTR199610, GTR199609, GTR199608, GTR199607, GTR199606, GTR199605, GTR199604, GTR199603, GTR199602, GTR199601, GTR199512, GTR199511, GTR199510, GTR199509, GTR199508, GTR199507, GTR199506, GTR199505, GTR199504, GTR199503, GTR199502, GTR199501, GTR199412, GTR199411, GTR199410, GTR199409, GTR199408, GTR199407, GTR199406"
sqlFormula = "INSERT INTO perf_gross (" + colum_names + ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

start_time = time.time()
print("- opening file...")
# ref: read Excel file: https://www.datacamp.com/community/tutorials/python-excel-tutorial
wb = load_workbook("./Perf_201809_Gross.xlsx")
# print(wb.get_sheet_names())
print("- opened in: [ %s ] seconds" % (time.time() - start_time))
start_time = time.time()

sheet = wb.get_sheet_by_name('Gross')
print("-- reading sheet: %s" % sheet.title)

for r in range(2, sheet.max_row + 1): # (2, sheet.max_row + 1), row and col in Excel starts at 1
	Name = sheet.cell(r, 1 ).value
	SecId = sheet.cell(r, 2 ).value
	Category = sheet.cell(r, 3 ).value
	GTR201809 = sheet.cell(r, 4 ).value
	GTR201808 = sheet.cell(r, 5 ).value
	GTR201807 = sheet.cell(r, 6 ).value
	GTR201806 = sheet.cell(r, 7 ).value
	GTR201805 = sheet.cell(r, 8 ).value
	GTR201804 = sheet.cell(r, 9 ).value
	GTR201803 = sheet.cell(r, 10 ).value
	GTR201802 = sheet.cell(r, 11 ).value
	GTR201801 = sheet.cell(r, 12 ).value
	GTR201712 = sheet.cell(r, 13 ).value
	GTR201711 = sheet.cell(r, 14 ).value
	GTR201710 = sheet.cell(r, 15 ).value
	GTR201709 = sheet.cell(r, 16 ).value
	GTR201708 = sheet.cell(r, 17 ).value
	GTR201707 = sheet.cell(r, 18 ).value
	GTR201706 = sheet.cell(r, 19 ).value
	GTR201705 = sheet.cell(r, 20 ).value
	GTR201704 = sheet.cell(r, 21 ).value
	GTR201703 = sheet.cell(r, 22 ).value
	GTR201702 = sheet.cell(r, 23 ).value
	GTR201701 = sheet.cell(r, 24 ).value
	GTR201612 = sheet.cell(r, 25 ).value
	GTR201611 = sheet.cell(r, 26 ).value
	GTR201610 = sheet.cell(r, 27 ).value
	GTR201609 = sheet.cell(r, 28 ).value
	GTR201608 = sheet.cell(r, 29 ).value
	GTR201607 = sheet.cell(r, 30 ).value
	GTR201606 = sheet.cell(r, 31 ).value
	GTR201605 = sheet.cell(r, 32 ).value
	GTR201604 = sheet.cell(r, 33 ).value
	GTR201603 = sheet.cell(r, 34 ).value
	GTR201602 = sheet.cell(r, 35 ).value
	GTR201601 = sheet.cell(r, 36 ).value
	GTR201512 = sheet.cell(r, 37 ).value
	GTR201511 = sheet.cell(r, 38 ).value
	GTR201510 = sheet.cell(r, 39 ).value
	GTR201509 = sheet.cell(r, 40 ).value
	GTR201508 = sheet.cell(r, 41 ).value
	GTR201507 = sheet.cell(r, 42 ).value
	GTR201506 = sheet.cell(r, 43 ).value
	GTR201505 = sheet.cell(r, 44 ).value
	GTR201504 = sheet.cell(r, 45 ).value
	GTR201503 = sheet.cell(r, 46 ).value
	GTR201502 = sheet.cell(r, 47 ).value
	GTR201501 = sheet.cell(r, 48 ).value
	GTR201412 = sheet.cell(r, 49 ).value
	GTR201411 = sheet.cell(r, 50 ).value
	GTR201410 = sheet.cell(r, 51 ).value
	GTR201409 = sheet.cell(r, 52 ).value
	GTR201408 = sheet.cell(r, 53 ).value
	GTR201407 = sheet.cell(r, 54 ).value
	GTR201406 = sheet.cell(r, 55 ).value
	GTR201405 = sheet.cell(r, 56 ).value
	GTR201404 = sheet.cell(r, 57 ).value
	GTR201403 = sheet.cell(r, 58 ).value
	GTR201402 = sheet.cell(r, 59 ).value
	GTR201401 = sheet.cell(r, 60 ).value
	GTR201312 = sheet.cell(r, 61 ).value
	GTR201311 = sheet.cell(r, 62 ).value
	GTR201310 = sheet.cell(r, 63 ).value
	GTR201309 = sheet.cell(r, 64 ).value
	GTR201308 = sheet.cell(r, 65 ).value
	GTR201307 = sheet.cell(r, 66 ).value
	GTR201306 = sheet.cell(r, 67 ).value
	GTR201305 = sheet.cell(r, 68 ).value
	GTR201304 = sheet.cell(r, 69 ).value
	GTR201303 = sheet.cell(r, 70 ).value
	GTR201302 = sheet.cell(r, 71 ).value
	GTR201301 = sheet.cell(r, 72 ).value
	GTR201212 = sheet.cell(r, 73 ).value
	GTR201211 = sheet.cell(r, 74 ).value
	GTR201210 = sheet.cell(r, 75 ).value
	GTR201209 = sheet.cell(r, 76 ).value
	GTR201208 = sheet.cell(r, 77 ).value
	GTR201207 = sheet.cell(r, 78 ).value
	GTR201206 = sheet.cell(r, 79 ).value
	GTR201205 = sheet.cell(r, 80 ).value
	GTR201204 = sheet.cell(r, 81 ).value
	GTR201203 = sheet.cell(r, 82 ).value
	GTR201202 = sheet.cell(r, 83 ).value
	GTR201201 = sheet.cell(r, 84 ).value
	GTR201112 = sheet.cell(r, 85 ).value
	GTR201111 = sheet.cell(r, 86 ).value
	GTR201110 = sheet.cell(r, 87 ).value
	GTR201109 = sheet.cell(r, 88 ).value
	GTR201108 = sheet.cell(r, 89 ).value
	GTR201107 = sheet.cell(r, 90 ).value
	GTR201106 = sheet.cell(r, 91 ).value
	GTR201105 = sheet.cell(r, 92 ).value
	GTR201104 = sheet.cell(r, 93 ).value
	GTR201103 = sheet.cell(r, 94 ).value
	GTR201102 = sheet.cell(r, 95 ).value
	GTR201101 = sheet.cell(r, 96 ).value
	GTR201012 = sheet.cell(r, 97 ).value
	GTR201011 = sheet.cell(r, 98 ).value
	GTR201010 = sheet.cell(r, 99 ).value
	GTR201009 = sheet.cell(r, 100 ).value
	GTR201008 = sheet.cell(r, 101 ).value
	GTR201007 = sheet.cell(r, 102 ).value
	GTR201006 = sheet.cell(r, 103 ).value
	GTR201005 = sheet.cell(r, 104 ).value
	GTR201004 = sheet.cell(r, 105 ).value
	GTR201003 = sheet.cell(r, 106 ).value
	GTR201002 = sheet.cell(r, 107 ).value
	GTR201001 = sheet.cell(r, 108 ).value
	GTR200912 = sheet.cell(r, 109 ).value
	GTR200911 = sheet.cell(r, 110 ).value
	GTR200910 = sheet.cell(r, 111 ).value
	GTR200909 = sheet.cell(r, 112 ).value
	GTR200908 = sheet.cell(r, 113 ).value
	GTR200907 = sheet.cell(r, 114 ).value
	GTR200906 = sheet.cell(r, 115 ).value
	GTR200905 = sheet.cell(r, 116 ).value
	GTR200904 = sheet.cell(r, 117 ).value
	GTR200903 = sheet.cell(r, 118 ).value
	GTR200902 = sheet.cell(r, 119 ).value
	GTR200901 = sheet.cell(r, 120 ).value
	GTR200812 = sheet.cell(r, 121 ).value
	GTR200811 = sheet.cell(r, 122 ).value
	GTR200810 = sheet.cell(r, 123 ).value
	GTR200809 = sheet.cell(r, 124 ).value
	GTR200808 = sheet.cell(r, 125 ).value
	GTR200807 = sheet.cell(r, 126 ).value
	GTR200806 = sheet.cell(r, 127 ).value
	GTR200805 = sheet.cell(r, 128 ).value
	GTR200804 = sheet.cell(r, 129 ).value
	GTR200803 = sheet.cell(r, 130 ).value
	GTR200802 = sheet.cell(r, 131 ).value
	GTR200801 = sheet.cell(r, 132 ).value
	GTR200712 = sheet.cell(r, 133 ).value
	GTR200711 = sheet.cell(r, 134 ).value
	GTR200710 = sheet.cell(r, 135 ).value
	GTR200709 = sheet.cell(r, 136 ).value
	GTR200708 = sheet.cell(r, 137 ).value
	GTR200707 = sheet.cell(r, 138 ).value
	GTR200706 = sheet.cell(r, 139 ).value
	GTR200705 = sheet.cell(r, 140 ).value
	GTR200704 = sheet.cell(r, 141 ).value
	GTR200703 = sheet.cell(r, 142 ).value
	GTR200702 = sheet.cell(r, 143 ).value
	GTR200701 = sheet.cell(r, 144 ).value
	GTR200612 = sheet.cell(r, 145 ).value
	GTR200611 = sheet.cell(r, 146 ).value
	GTR200610 = sheet.cell(r, 147 ).value
	GTR200609 = sheet.cell(r, 148 ).value
	GTR200608 = sheet.cell(r, 149 ).value
	GTR200607 = sheet.cell(r, 150 ).value
	GTR200606 = sheet.cell(r, 151 ).value
	GTR200605 = sheet.cell(r, 152 ).value
	GTR200604 = sheet.cell(r, 153 ).value
	GTR200603 = sheet.cell(r, 154 ).value
	GTR200602 = sheet.cell(r, 155 ).value
	GTR200601 = sheet.cell(r, 156 ).value
	GTR200512 = sheet.cell(r, 157 ).value
	GTR200511 = sheet.cell(r, 158 ).value
	GTR200510 = sheet.cell(r, 159 ).value
	GTR200509 = sheet.cell(r, 160 ).value
	GTR200508 = sheet.cell(r, 161 ).value
	GTR200507 = sheet.cell(r, 162 ).value
	GTR200506 = sheet.cell(r, 163 ).value
	GTR200505 = sheet.cell(r, 164 ).value
	GTR200504 = sheet.cell(r, 165 ).value
	GTR200503 = sheet.cell(r, 166 ).value
	GTR200502 = sheet.cell(r, 167 ).value
	GTR200501 = sheet.cell(r, 168 ).value
	GTR200412 = sheet.cell(r, 169 ).value
	GTR200411 = sheet.cell(r, 170 ).value
	GTR200410 = sheet.cell(r, 171 ).value
	GTR200409 = sheet.cell(r, 172 ).value
	GTR200408 = sheet.cell(r, 173 ).value
	GTR200407 = sheet.cell(r, 174 ).value
	GTR200406 = sheet.cell(r, 175 ).value
	GTR200405 = sheet.cell(r, 176 ).value
	GTR200404 = sheet.cell(r, 177 ).value
	GTR200403 = sheet.cell(r, 178 ).value
	GTR200402 = sheet.cell(r, 179 ).value
	GTR200401 = sheet.cell(r, 180 ).value
	GTR200312 = sheet.cell(r, 181 ).value
	GTR200311 = sheet.cell(r, 182 ).value
	GTR200310 = sheet.cell(r, 183 ).value
	GTR200309 = sheet.cell(r, 184 ).value
	GTR200308 = sheet.cell(r, 185 ).value
	GTR200307 = sheet.cell(r, 186 ).value
	GTR200306 = sheet.cell(r, 187 ).value
	GTR200305 = sheet.cell(r, 188 ).value
	GTR200304 = sheet.cell(r, 189 ).value
	GTR200303 = sheet.cell(r, 190 ).value
	GTR200302 = sheet.cell(r, 191 ).value
	GTR200301 = sheet.cell(r, 192 ).value
	GTR200212 = sheet.cell(r, 193 ).value
	GTR200211 = sheet.cell(r, 194 ).value
	GTR200210 = sheet.cell(r, 195 ).value
	GTR200209 = sheet.cell(r, 196 ).value
	GTR200208 = sheet.cell(r, 197 ).value
	GTR200207 = sheet.cell(r, 198 ).value
	GTR200206 = sheet.cell(r, 199 ).value
	GTR200205 = sheet.cell(r, 200 ).value
	GTR200204 = sheet.cell(r, 201 ).value
	GTR200203 = sheet.cell(r, 202 ).value
	GTR200202 = sheet.cell(r, 203 ).value
	GTR200201 = sheet.cell(r, 204 ).value
	GTR200112 = sheet.cell(r, 205 ).value
	GTR200111 = sheet.cell(r, 206 ).value
	GTR200110 = sheet.cell(r, 207 ).value
	GTR200109 = sheet.cell(r, 208 ).value
	GTR200108 = sheet.cell(r, 209 ).value
	GTR200107 = sheet.cell(r, 210 ).value
	GTR200106 = sheet.cell(r, 211 ).value
	GTR200105 = sheet.cell(r, 212 ).value
	GTR200104 = sheet.cell(r, 213 ).value
	GTR200103 = sheet.cell(r, 214 ).value
	GTR200102 = sheet.cell(r, 215 ).value
	GTR200101 = sheet.cell(r, 216 ).value
	GTR200012 = sheet.cell(r, 217 ).value
	GTR200011 = sheet.cell(r, 218 ).value
	GTR200010 = sheet.cell(r, 219 ).value
	GTR200009 = sheet.cell(r, 220 ).value
	GTR200008 = sheet.cell(r, 221 ).value
	GTR200007 = sheet.cell(r, 222 ).value
	GTR200006 = sheet.cell(r, 223 ).value
	GTR200005 = sheet.cell(r, 224 ).value
	GTR200004 = sheet.cell(r, 225 ).value
	GTR200003 = sheet.cell(r, 226 ).value
	GTR200002 = sheet.cell(r, 227 ).value
	GTR200001 = sheet.cell(r, 228 ).value
	GTR199912 = sheet.cell(r, 229 ).value
	GTR199911 = sheet.cell(r, 230 ).value
	GTR199910 = sheet.cell(r, 231 ).value
	GTR199909 = sheet.cell(r, 232 ).value
	GTR199908 = sheet.cell(r, 233 ).value
	GTR199907 = sheet.cell(r, 234 ).value
	GTR199906 = sheet.cell(r, 235 ).value
	GTR199905 = sheet.cell(r, 236 ).value
	GTR199904 = sheet.cell(r, 237 ).value
	GTR199903 = sheet.cell(r, 238 ).value
	GTR199902 = sheet.cell(r, 239 ).value
	GTR199901 = sheet.cell(r, 240 ).value
	GTR199812 = sheet.cell(r, 241 ).value
	GTR199811 = sheet.cell(r, 242 ).value
	GTR199810 = sheet.cell(r, 243 ).value
	GTR199809 = sheet.cell(r, 244 ).value
	GTR199808 = sheet.cell(r, 245 ).value
	GTR199807 = sheet.cell(r, 246 ).value
	GTR199806 = sheet.cell(r, 247 ).value
	GTR199805 = sheet.cell(r, 248 ).value
	GTR199804 = sheet.cell(r, 249 ).value
	GTR199803 = sheet.cell(r, 250 ).value
	GTR199802 = sheet.cell(r, 251 ).value
	GTR199801 = sheet.cell(r, 252 ).value
	GTR199712 = sheet.cell(r, 253 ).value
	GTR199711 = sheet.cell(r, 254 ).value
	GTR199710 = sheet.cell(r, 255 ).value
	GTR199709 = sheet.cell(r, 256 ).value
	GTR199708 = sheet.cell(r, 257 ).value
	GTR199707 = sheet.cell(r, 258 ).value
	GTR199706 = sheet.cell(r, 259 ).value
	GTR199705 = sheet.cell(r, 260 ).value
	GTR199704 = sheet.cell(r, 261 ).value
	GTR199703 = sheet.cell(r, 262 ).value
	GTR199702 = sheet.cell(r, 263 ).value
	GTR199701 = sheet.cell(r, 264 ).value
	GTR199612 = sheet.cell(r, 265 ).value
	GTR199611 = sheet.cell(r, 266 ).value
	GTR199610 = sheet.cell(r, 267 ).value
	GTR199609 = sheet.cell(r, 268 ).value
	GTR199608 = sheet.cell(r, 269 ).value
	GTR199607 = sheet.cell(r, 270 ).value
	GTR199606 = sheet.cell(r, 271 ).value
	GTR199605 = sheet.cell(r, 272 ).value
	GTR199604 = sheet.cell(r, 273 ).value
	GTR199603 = sheet.cell(r, 274 ).value
	GTR199602 = sheet.cell(r, 275 ).value
	GTR199601 = sheet.cell(r, 276 ).value
	GTR199512 = sheet.cell(r, 277 ).value
	GTR199511 = sheet.cell(r, 278 ).value
	GTR199510 = sheet.cell(r, 279 ).value
	GTR199509 = sheet.cell(r, 280 ).value
	GTR199508 = sheet.cell(r, 281 ).value
	GTR199507 = sheet.cell(r, 282 ).value
	GTR199506 = sheet.cell(r, 283 ).value
	GTR199505 = sheet.cell(r, 284 ).value
	GTR199504 = sheet.cell(r, 285 ).value
	GTR199503 = sheet.cell(r, 286 ).value
	GTR199502 = sheet.cell(r, 287 ).value
	GTR199501 = sheet.cell(r, 288 ).value
	GTR199412 = sheet.cell(r, 289 ).value
	GTR199411 = sheet.cell(r, 290 ).value
	GTR199410 = sheet.cell(r, 291 ).value
	GTR199409 = sheet.cell(r, 292 ).value
	GTR199408 = sheet.cell(r, 293 ).value
	GTR199407 = sheet.cell(r, 294 ).value
	GTR199406 = sheet.cell(r, 295 ).value
	
	values = (Name, SecId, Category, GTR201809, GTR201808, GTR201807, GTR201806, GTR201805, GTR201804, GTR201803, GTR201802, GTR201801, GTR201712, GTR201711, GTR201710, GTR201709, GTR201708, GTR201707, GTR201706, GTR201705, GTR201704, GTR201703, GTR201702, GTR201701, GTR201612, GTR201611, GTR201610, GTR201609, GTR201608, GTR201607, GTR201606, GTR201605, GTR201604, GTR201603, GTR201602, GTR201601, GTR201512, GTR201511, GTR201510, GTR201509, GTR201508, GTR201507, GTR201506, GTR201505, GTR201504, GTR201503, GTR201502, GTR201501, GTR201412, GTR201411, GTR201410, GTR201409, GTR201408, GTR201407, GTR201406, GTR201405, GTR201404, GTR201403, GTR201402, GTR201401, GTR201312, GTR201311, GTR201310, GTR201309, GTR201308, GTR201307, GTR201306, GTR201305, GTR201304, GTR201303, GTR201302, GTR201301, GTR201212, GTR201211, GTR201210, GTR201209, GTR201208, GTR201207, GTR201206, GTR201205, GTR201204, GTR201203, GTR201202, GTR201201, GTR201112, GTR201111, GTR201110, GTR201109, GTR201108, GTR201107, GTR201106, GTR201105, GTR201104, GTR201103, GTR201102, GTR201101, GTR201012, GTR201011, GTR201010, GTR201009, GTR201008, GTR201007, GTR201006, GTR201005, GTR201004, GTR201003, GTR201002, GTR201001, GTR200912, GTR200911, GTR200910, GTR200909, GTR200908, GTR200907, GTR200906, GTR200905, GTR200904, GTR200903, GTR200902, GTR200901, GTR200812, GTR200811, GTR200810, GTR200809, GTR200808, GTR200807, GTR200806, GTR200805, GTR200804, GTR200803, GTR200802, GTR200801, GTR200712, GTR200711, GTR200710, GTR200709, GTR200708, GTR200707, GTR200706, GTR200705, GTR200704, GTR200703, GTR200702, GTR200701, GTR200612, GTR200611, GTR200610, GTR200609, GTR200608, GTR200607, GTR200606, GTR200605, GTR200604, GTR200603, GTR200602, GTR200601, GTR200512, GTR200511, GTR200510, GTR200509, GTR200508, GTR200507, GTR200506, GTR200505, GTR200504, GTR200503, GTR200502, GTR200501, GTR200412, GTR200411, GTR200410, GTR200409, GTR200408, GTR200407, GTR200406, GTR200405, GTR200404, GTR200403, GTR200402, GTR200401, GTR200312, GTR200311, GTR200310, GTR200309, GTR200308, GTR200307, GTR200306, GTR200305, GTR200304, GTR200303, GTR200302, GTR200301, GTR200212, GTR200211, GTR200210, GTR200209, GTR200208, GTR200207, GTR200206, GTR200205, GTR200204, GTR200203, GTR200202, GTR200201, GTR200112, GTR200111, GTR200110, GTR200109, GTR200108, GTR200107, GTR200106, GTR200105, GTR200104, GTR200103, GTR200102, GTR200101, GTR200012, GTR200011, GTR200010, GTR200009, GTR200008, GTR200007, GTR200006, GTR200005, GTR200004, GTR200003, GTR200002, GTR200001, GTR199912, GTR199911, GTR199910, GTR199909, GTR199908, GTR199907, GTR199906, GTR199905, GTR199904, GTR199903, GTR199902, GTR199901, GTR199812, GTR199811, GTR199810, GTR199809, GTR199808, GTR199807, GTR199806, GTR199805, GTR199804, GTR199803, GTR199802, GTR199801, GTR199712, GTR199711, GTR199710, GTR199709, GTR199708, GTR199707, GTR199706, GTR199705, GTR199704, GTR199703, GTR199702, GTR199701, GTR199612, GTR199611, GTR199610, GTR199609, GTR199608, GTR199607, GTR199606, GTR199605, GTR199604, GTR199603, GTR199602, GTR199601, GTR199512, GTR199511, GTR199510, GTR199509, GTR199508, GTR199507, GTR199506, GTR199505, GTR199504, GTR199503, GTR199502, GTR199501, GTR199412, GTR199411, GTR199410, GTR199409, GTR199408, GTR199407, GTR199406)

	# print(r)
	# print("--- inserting val = %s", values)
	mycursor.execute(sqlFormula, values)

print("--- finished inserting in: [ %s ] seconds" % (time.time() - start_time))
print("--- close db cursor.")
mycursor.close()

print("--- will commit to db...")
mydb.commit()

print("--- close db.")
mydb.close()



import mysql.connector
import time
from openpyxl import load_workbook
import db_config


mydb = mysql.connector.connect(
	host=db_config.local_host,
	user=db_config.user,
	passwd=db_config.passwd,
	database=db_config.database	
	)

mycursor = mydb.cursor()

colum_names = "GR_ID, FundId, SecId, MSCat, ProdType, MFVA, Fund, Advisor, VIT, MultiSub, SubAdv, SubAdvd, RealAdv, SubSch, SubStart, SubRate, SubAlloc, MgdVol, SmartBet, WatchLst, InceptDt, InceptYr, MgrName, MgrTenL, MgrTenA, PrBench, MgtFee, CashP, TurnRate, EnhIndex, Indexed, FoF, ESG, NonDiv, CloseAll, CloseNew, MastFeed, Sleeve, TeamMgd, Bench, Bench_ID, Yld12M, SECYld, AS201809, AS201712, AS201612, AS201512, AS201412, AS201312, AS201212, AS201112, AS201012, AS200912, AS200812, AS200712, AS200612, AS200512, AS200412, AS200312, NFYTD, NF3M, MF6M, NF1Y, NF2Y, NF3Y, NF4Y, NF5Y, NF10Y, NF15Y, NF2017, NF2016, NF2015, NF2014, NF2013, NF2012, NF2011, NF2010, NF2009, NF2008, NF2007, NF2006, NF2005, NF2004, NF2003, QKYTD, QK3M, QK6M, QK9M, QK1Y, QK2Y, QK3Y, QK4Y, QK5Y, QK10Y, QK15Y, QK2017, QK2016, QK2015, QK2014, QK2013, QK2012, QK2011, QK2010, QK2009, QK2008, QKAlph3, QKExRt3, QKShrp3, QKInfR3, QKBeta3, QKStdv3, QKRsq3, QKUpS3, QKDnS3, QKTrEr3, QKFUps3, QKFDns3, QKBUps3, QKBDns3, QKBAvg3, QKOutp3, APAUM, AP201809, AP201712, AP201612, AP201512, AP201412, AP201312, AP201212, AP201112, AP201012, AP200912, AP200812, AP200712, AP200612, AP200512, AP200412, AP200312, NFPYTD, NFP3M, NFP6M, NFP1Y, NFP2Y, NFP3Y, NFP4Y, NFP5Y, NFP10Y, NFP15Y, NFP2017, NFP2016, NFP2015, NFP2014, NFP2013, NFP2012, NFP2011, NFP2010, NFP2009, NFP2008, NFP2007, NFP2006, NFP2005, NFP2004, NFP2003, SubAdvP, SubAdvdP, SubSchP, SubStrtP, SubRateP, Platform, Platcnt, valgro2, smlrg2, valgro3, smlrg3, valgro5, smlrg5, AdvPar, SAdvPar, MandID, SubDollr"
sqlFormula = "INSERT INTO blog_Funds (" + colum_names + ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

start_time = time.time()
print("- opening file...")
# ref: read Excel file: https://www.datacamp.com/community/tutorials/python-excel-tutorial
wb = load_workbook("./Fund_Data.xlsx")
# print(wb.get_sheet_names())
print("- opened in: [ %s ] seconds" % (time.time() - start_time))
start_time = time.time()

sheet = wb.get_sheet_by_name('Fund')
print("-- reading sheet: %s" % sheet.title)

for r in range(8, sheet.max_row + 1): # (2, sheet.max_row + 1), row and col in Excel starts at 1
	GR_ID	= sheet.cell(r,	2).value
	FundId	= sheet.cell(r,	3).value
	SecId	= sheet.cell(r,	4).value
	MSCat	= sheet.cell(r,	5).value
	ProdType	= sheet.cell(r,	6).value
	MFVA	= sheet.cell(r,	7).value
	Fund	= sheet.cell(r,	8).value
	Advisor	= sheet.cell(r,	9).value
	VIT	= sheet.cell(r,	10).value
	MultiSub	= sheet.cell(r,	11).value
	SubAdv	= sheet.cell(r,	12).value
	SubAdvd	= sheet.cell(r,	13).value
	RealAdv	= sheet.cell(r,	14).value
	SubSch	= sheet.cell(r,	15).value
	SubStart	= sheet.cell(r,	16).value
	SubRate	= sheet.cell(r,	17).value
	SubAlloc	= sheet.cell(r,	18).value
	MgdVol	= sheet.cell(r,	19).value
	SmartBet	= sheet.cell(r,	20).value
	WatchLst	= sheet.cell(r,	21).value
	InceptDt	= sheet.cell(r,	22).value
	InceptYr	= sheet.cell(r,	23).value
	MgrName	= sheet.cell(r,	24).value
	MgrTenL	= sheet.cell(r,	25).value
	MgrTenA	= sheet.cell(r,	26).value
	PrBench	= sheet.cell(r,	27).value
	MgtFee	= sheet.cell(r,	28).value
	CashP	= sheet.cell(r,	29).value
	TurnRate	= sheet.cell(r,	30).value
	EnhIndex	= sheet.cell(r,	31).value
	Indexed	= sheet.cell(r,	32).value
	FoF	= sheet.cell(r,	33).value
	ESG	= sheet.cell(r,	34).value
	NonDiv	= sheet.cell(r,	35).value
	CloseAll	= sheet.cell(r,	36).value
	CloseNew	= sheet.cell(r,	37).value
	MastFeed	= sheet.cell(r,	38).value
	Sleeve	= sheet.cell(r,	39).value
	TeamMgd	= sheet.cell(r,	40).value
	Bench	= sheet.cell(r,	41).value
	Bench_ID	= sheet.cell(r,	42).value
	Yld12M	= sheet.cell(r,	43).value
	SECYld	= sheet.cell(r,	44).value
	AS201809	= sheet.cell(r,	45).value
	AS201712	= sheet.cell(r,	46).value
	AS201612	= sheet.cell(r,	47).value
	AS201512	= sheet.cell(r,	48).value
	AS201412	= sheet.cell(r,	49).value
	AS201312	= sheet.cell(r,	50).value
	AS201212	= sheet.cell(r,	51).value
	AS201112	= sheet.cell(r,	52).value
	AS201012	= sheet.cell(r,	53).value
	AS200912	= sheet.cell(r,	54).value
	AS200812	= sheet.cell(r,	55).value
	AS200712	= sheet.cell(r,	56).value
	AS200612	= sheet.cell(r,	57).value
	AS200512	= sheet.cell(r,	58).value
	AS200412	= sheet.cell(r,	59).value
	AS200312	= sheet.cell(r,	60).value
	NFYTD	= sheet.cell(r,	61).value
	NF3M	= sheet.cell(r,	62).value
	MF6M	= sheet.cell(r,	63).value
	NF1Y	= sheet.cell(r,	64).value
	NF2Y	= sheet.cell(r,	65).value
	NF3Y	= sheet.cell(r,	66).value
	NF4Y	= sheet.cell(r,	67).value
	NF5Y	= sheet.cell(r,	68).value
	NF10Y	= sheet.cell(r,	69).value
	NF15Y	= sheet.cell(r,	70).value
	NF2017	= sheet.cell(r,	71).value
	NF2016	= sheet.cell(r,	72).value
	NF2015	= sheet.cell(r,	73).value
	NF2014	= sheet.cell(r,	74).value
	NF2013	= sheet.cell(r,	75).value
	NF2012	= sheet.cell(r,	76).value
	NF2011	= sheet.cell(r,	77).value
	NF2010	= sheet.cell(r,	78).value
	NF2009	= sheet.cell(r,	79).value
	NF2008	= sheet.cell(r,	80).value
	NF2007	= sheet.cell(r,	81).value
	NF2006	= sheet.cell(r,	82).value
	NF2005	= sheet.cell(r,	83).value
	NF2004	= sheet.cell(r,	84).value
	NF2003	= sheet.cell(r,	85).value
	QKYTD	= sheet.cell(r,	86).value
	QK3M	= sheet.cell(r,	87).value
	QK6M	= sheet.cell(r,	88).value
	QK9M	= sheet.cell(r,	89).value
	QK1Y	= sheet.cell(r,	90).value
	QK2Y	= sheet.cell(r,	91).value
	QK3Y	= sheet.cell(r,	92).value
	QK4Y	= sheet.cell(r,	93).value
	QK5Y	= sheet.cell(r,	94).value
	QK10Y	= sheet.cell(r,	95).value
	QK15Y	= sheet.cell(r,	96).value
	QK2017	= sheet.cell(r,	97).value
	QK2016	= sheet.cell(r,	98).value
	QK2015	= sheet.cell(r,	99).value
	QK2014	= sheet.cell(r,	100).value
	QK2013	= sheet.cell(r,	101).value
	QK2012	= sheet.cell(r,	102).value
	QK2011	= sheet.cell(r,	103).value
	QK2010	= sheet.cell(r,	104).value
	QK2009	= sheet.cell(r,	105).value
	QK2008	= sheet.cell(r,	106).value
	QKAlph3	= sheet.cell(r,	107).value
	QKExRt3	= sheet.cell(r,	108).value
	QKShrp3	= sheet.cell(r,	109).value
	QKInfR3	= sheet.cell(r,	110).value
	QKBeta3	= sheet.cell(r,	111).value
	QKStdv3	= sheet.cell(r,	112).value
	QKRsq3	= sheet.cell(r,	113).value
	QKUpS3	= sheet.cell(r,	114).value
	QKDnS3	= sheet.cell(r,	115).value
	QKTrEr3	= sheet.cell(r,	116).value
	QKFUps3	= sheet.cell(r,	117).value
	QKFDns3	= sheet.cell(r,	118).value
	QKBUps3	= sheet.cell(r,	119).value
	QKBDns3	= sheet.cell(r,	120).value
	QKBAvg3	= sheet.cell(r,	121).value
	QKOutp3	= sheet.cell(r,	122).value
	APAUM	= sheet.cell(r,	123).value
	AP201809	= sheet.cell(r,	124).value
	AP201712	= sheet.cell(r,	125).value
	AP201612	= sheet.cell(r,	126).value
	AP201512	= sheet.cell(r,	127).value
	AP201412	= sheet.cell(r,	128).value
	AP201312	= sheet.cell(r,	129).value
	AP201212	= sheet.cell(r,	130).value
	AP201112	= sheet.cell(r,	131).value
	AP201012	= sheet.cell(r,	132).value
	AP200912	= sheet.cell(r,	133).value
	AP200812	= sheet.cell(r,	134).value
	AP200712	= sheet.cell(r,	135).value
	AP200612	= sheet.cell(r,	136).value
	AP200512	= sheet.cell(r,	137).value
	AP200412	= sheet.cell(r,	138).value
	AP200312	= sheet.cell(r,	139).value
	NFPYTD	= sheet.cell(r,	140).value
	NFP3M	= sheet.cell(r,	141).value
	NFP6M	= sheet.cell(r,	142).value
	NFP1Y	= sheet.cell(r,	143).value
	NFP2Y	= sheet.cell(r,	144).value
	NFP3Y	= sheet.cell(r,	145).value
	NFP4Y	= sheet.cell(r,	146).value
	NFP5Y	= sheet.cell(r,	147).value
	NFP10Y	= sheet.cell(r,	148).value
	NFP15Y	= sheet.cell(r,	149).value
	NFP2017	= sheet.cell(r,	150).value
	NFP2016	= sheet.cell(r,	151).value
	NFP2015	= sheet.cell(r,	152).value
	NFP2014	= sheet.cell(r,	153).value
	NFP2013	= sheet.cell(r,	154).value
	NFP2012	= sheet.cell(r,	155).value
	NFP2011	= sheet.cell(r,	156).value
	NFP2010	= sheet.cell(r,	157).value
	NFP2009	= sheet.cell(r,	158).value
	NFP2008	= sheet.cell(r,	159).value
	NFP2007	= sheet.cell(r,	160).value
	NFP2006	= sheet.cell(r,	161).value
	NFP2005	= sheet.cell(r,	162).value
	NFP2004	= sheet.cell(r,	163).value
	NFP2003	= sheet.cell(r,	164).value
	SubAdvP	= sheet.cell(r,	165).value
	SubAdvdP	= sheet.cell(r,	166).value
	SubSchP	= sheet.cell(r,	167).value
	SubStrtP	= sheet.cell(r,	168).value
	SubRateP	= sheet.cell(r,	169).value
	Platform	= sheet.cell(r,	170).value
	Platcnt	= sheet.cell(r,	171).value
	valgro2	= sheet.cell(r,	172).value
	smlrg2	= sheet.cell(r,	173).value
	valgro3	= sheet.cell(r,	174).value
	smlrg3	= sheet.cell(r,	175).value
	valgro5	= sheet.cell(r,	176).value
	smlrg5	= sheet.cell(r,	177).value
	AdvPar	= sheet.cell(r,	178).value
	SAdvPar	= sheet.cell(r,	179).value
	MandID	= sheet.cell(r,	180).value
	SubDollr	= sheet.cell(r,	181).value

	values = (GR_ID, FundId, SecId, MSCat, ProdType, MFVA, Fund, Advisor, VIT, MultiSub, SubAdv, SubAdvd, RealAdv, SubSch, SubStart, SubRate, SubAlloc, MgdVol, SmartBet, WatchLst, InceptDt, InceptYr, MgrName, MgrTenL, MgrTenA, PrBench, MgtFee, CashP, TurnRate, EnhIndex, Indexed, FoF, ESG, NonDiv, CloseAll, CloseNew, MastFeed, Sleeve, TeamMgd, Bench, Bench_ID, Yld12M, SECYld, AS201809, AS201712, AS201612, AS201512, AS201412, AS201312, AS201212, AS201112, AS201012, AS200912, AS200812, AS200712, AS200612, AS200512, AS200412, AS200312, NFYTD, NF3M, MF6M, NF1Y, NF2Y, NF3Y, NF4Y, NF5Y, NF10Y, NF15Y, NF2017, NF2016, NF2015, NF2014, NF2013, NF2012, NF2011, NF2010, NF2009, NF2008, NF2007, NF2006, NF2005, NF2004, NF2003, QKYTD, QK3M, QK6M, QK9M, QK1Y, QK2Y, QK3Y, QK4Y, QK5Y, QK10Y, QK15Y, QK2017, QK2016, QK2015, QK2014, QK2013, QK2012, QK2011, QK2010, QK2009, QK2008, QKAlph3, QKExRt3, QKShrp3, QKInfR3, QKBeta3, QKStdv3, QKRsq3, QKUpS3, QKDnS3, QKTrEr3, QKFUps3, QKFDns3, QKBUps3, QKBDns3, QKBAvg3, QKOutp3, APAUM, AP201809, AP201712, AP201612, AP201512, AP201412, AP201312, AP201212, AP201112, AP201012, AP200912, AP200812, AP200712, AP200612, AP200512, AP200412, AP200312, NFPYTD, NFP3M, NFP6M, NFP1Y, NFP2Y, NFP3Y, NFP4Y, NFP5Y, NFP10Y, NFP15Y, NFP2017, NFP2016, NFP2015, NFP2014, NFP2013, NFP2012, NFP2011, NFP2010, NFP2009, NFP2008, NFP2007, NFP2006, NFP2005, NFP2004, NFP2003, SubAdvP, SubAdvdP, SubSchP, SubStrtP, SubRateP, Platform, Platcnt, valgro2, smlrg2, valgro3, smlrg3, valgro5, smlrg5, AdvPar, SAdvPar, MandID, SubDollr)

	# print(r)
	# print("--- inserting val = %s", values)
	mycursor.execute(sqlFormula, values)

print("--- finished inserting in: [ %s ] seconds" % (time.time() - start_time))
print("--- close db cursor.")
mycursor.close()

print("--- will commit to db...")
mydb.commit()

print("--- close db.")
mydb.close()



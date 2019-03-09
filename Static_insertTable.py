import mysql.connector
from openpyxl import load_workbook

mydb = mysql.connector.connect(
	host="127.0.0.1",
	user="root",
	passwd="BXY5201314",
	database="gene"
	)

mycursor = mydb.cursor()

colum_names = "Name, CUSIP, SecId, FundId, MRFundID, PerfID, Ticker, OldestSh, Advisor, FirmName, BrndName, BrndNmID, InceptDt, GlBrdGrp, GlBrdCat, USCatGrp, MSCat, MSCatID, MSCatStr, MSInsCat, MSRatAll, MSRat3Y, MSRat5Y, MSRat10Y, EqStyleL, EqStyleS, FxStyleL, MSSubAdv, MSSubFee, TeamMgd, MgrHist, MgrName, MgrTnrAv, MgrTnrLg, PPBnch, PPBnchID, PPBIncpt, SPBench, AUMShr, AUMShrDt, FSizeCur, FSizeDt, hldBondL, hldBondS, hldLong, hldOthrL, hldOthrS, hldStckL, hldStckS, hldT10, hldofInv, hldAT10, hldAT10D, Yld12M, Yld12MD, SECYld, SECYldD, DivdCur, DivdCurD, DivdNAVC, PortDate, VirtClss, FeeMgt, NExpSAR, NExpSARD, ChrgAR, ChrgARD, NExpPrs, NExpPrsD, ObjPros, Fee12b1, TurnRat, TurnRatD, CashN, AABondN, AACashN, AAEqN, AANUSBdN, AANUSEqN, AAOthrN, AAUSBdN, AAUSEqN, AAABN, AABondL, AACashL, AAEqL, AANUSBdL, AANUSEqL, AAOthrL, AAUSBdL, AAUSEqL, FoFIH, Leveragd, FeederFd, MasterFd, MasterN, MSTRFundID, Sharia, Ethical, ModelFd, Avail529, AvailRet, AvailIns, TaxMgd, Contrarn, BrkAvail, EnhIndex, Indexed, FoF, SocCons, NonDiv, ClsdAll, ClsdALLD, ClsdNew, ClsdNewD, FeeAccnt, FeeAdmin, FeeAdv, FeeAudit, FeeCust, FeeDistr, FeeIns, FeeLegal, FeeOrg, FeeOthr, FeePerf, FeeProf, FeeReg, FeeShRpt, FeeTA, Domicile, BaseCurr, ObsoltD, ObsoltTy, DiscCurr, ShrType, AABdLR, AABdL, AACshLR, AACBLR, AAEqLR, AAxUSBLR, AAxUSELR, AAOtLR, AAPSLR, AAUSBLR, AAUSELR"
sqlFormula = "INSERT INTO company (" + colum_names + ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

print("opening file...")
# ref: read Excel file: https://www.datacamp.com/community/tutorials/python-excel-tutorial
wb = load_workbook("./Static_Data.xlsx")
# print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('MF')
print("get sheet: %s" % sheet.title)

for r in range(2, sheet.max_row + 1): # (2, sheet.max_row + 1), row and col in Excel starts at 1
	Name	= sheet.cell(r,	1).value
	CUSIP	= sheet.cell(r,	2).value
	SecId	= sheet.cell(r,	3).value
	FundId	= sheet.cell(r,	4).value
	MRFundID	= sheet.cell(r,	5).value
	PerfID	= sheet.cell(r,	6).value
	Ticker	= sheet.cell(r,	7).value
	OldestSh	= sheet.cell(r,	8).value
	Advisor	= sheet.cell(r,	9).value
	FirmName	= sheet.cell(r,	10).value
	BrndName	= sheet.cell(r,	11).value
	BrndNmID	= sheet.cell(r,	12).value
	InceptDt	= sheet.cell(r,	13).value
	GlBrdGrp	= sheet.cell(r,	14).value
	GlBrdCat	= sheet.cell(r,	15).value
	USCatGrp	= sheet.cell(r,	16).value
	MSCat	= sheet.cell(r,	17).value
	MSCatID	= sheet.cell(r,	18).value
	MSCatStr	= sheet.cell(r,	19).value
	MSInsCat	= sheet.cell(r,	20).value
	MSRatAll	= sheet.cell(r,	21).value
	MSRat3Y	= sheet.cell(r,	22).value
	MSRat5Y	= sheet.cell(r,	23).value
	MSRat10Y	= sheet.cell(r,	24).value
	EqStyleL	= sheet.cell(r,	25).value
	EqStyleS	= sheet.cell(r,	26).value
	FxStyleL	= sheet.cell(r,	27).value
	MSSubAdv	= sheet.cell(r,	28).value
	MSSubFee	= sheet.cell(r,	29).value
	TeamMgd	= sheet.cell(r,	30).value
	MgrHist	= sheet.cell(r,	31).value
	MgrName	= sheet.cell(r,	32).value
	MgrTnrAv	= sheet.cell(r,	33).value
	MgrTnrLg	= sheet.cell(r,	34).value
	PPBnch	= sheet.cell(r,	35).value
	PPBnchID	= sheet.cell(r,	36).value
	PPBIncpt	= sheet.cell(r,	37).value
	SPBench	= sheet.cell(r,	38).value
	AUMShr	= sheet.cell(r,	39).value
	AUMShrDt	= sheet.cell(r,	40).value
	FSizeCur	= sheet.cell(r,	41).value
	FSizeDt	= sheet.cell(r,	42).value
	hldBondL	= sheet.cell(r,	43).value
	hldBondS	= sheet.cell(r,	44).value
	hldLong	= sheet.cell(r,	45).value
	hldOthrL	= sheet.cell(r,	46).value
	hldOthrS	= sheet.cell(r,	47).value
	hldStckL	= sheet.cell(r,	48).value
	hldStckS	= sheet.cell(r,	49).value
	hldT10	= sheet.cell(r,	50).value
	hldofInv	= sheet.cell(r,	51).value
	hldAT10	= sheet.cell(r,	52).value
	hldAT10D	= sheet.cell(r,	53).value
	Yld12M	= sheet.cell(r,	54).value
	Yld12MD	= sheet.cell(r,	55).value
	SECYld	= sheet.cell(r,	56).value
	SECYldD	= sheet.cell(r,	57).value
	DivdCur	= sheet.cell(r,	58).value
	DivdCurD	= sheet.cell(r,	59).value
	DivdNAVC	= sheet.cell(r,	60).value
	PortDate	= sheet.cell(r,	61).value
	VirtClss	= sheet.cell(r,	62).value
	FeeMgt	= sheet.cell(r,	63).value
	NExpSAR	= sheet.cell(r,	64).value
	NExpSARD	= sheet.cell(r,	65).value
	ChrgAR	= sheet.cell(r,	66).value
	ChrgARD	= sheet.cell(r,	67).value
	NExpPrs	= sheet.cell(r,	68).value
	NExpPrsD	= sheet.cell(r,	69).value
	ObjPros	= sheet.cell(r,	70).value
	Fee12b1	= sheet.cell(r,	71).value
	TurnRat	= sheet.cell(r,	72).value
	TurnRatD	= sheet.cell(r,	73).value
	CashN	= sheet.cell(r,	74).value
	AABondN	= sheet.cell(r,	75).value
	AACashN	= sheet.cell(r,	76).value
	AAEqN	= sheet.cell(r,	77).value
	AANUSBdN	= sheet.cell(r,	78).value
	AANUSEqN	= sheet.cell(r,	79).value
	AAOthrN	= sheet.cell(r,	80).value
	AAUSBdN	= sheet.cell(r,	81).value
	AAUSEqN	= sheet.cell(r,	82).value
	AAABN	= sheet.cell(r,	83).value
	AABondL	= sheet.cell(r,	84).value
	AACashL	= sheet.cell(r,	85).value
	AAEqL	= sheet.cell(r,	86).value
	AANUSBdL	= sheet.cell(r,	87).value
	AANUSEqL	= sheet.cell(r,	88).value
	AAOthrL	= sheet.cell(r,	89).value
	AAUSBdL	= sheet.cell(r,	90).value
	AAUSEqL	= sheet.cell(r,	91).value
	FoFIH	= sheet.cell(r,	92).value
	Leveragd	= sheet.cell(r,	93).value
	FeederFd	= sheet.cell(r,	94).value
	MasterFd	= sheet.cell(r,	95).value
	MasterN	= sheet.cell(r,	96).value
	MSTRFundID	= sheet.cell(r,	97).value
	Sharia	= sheet.cell(r,	98).value
	Ethical	= sheet.cell(r,	99).value
	ModelFd	= sheet.cell(r,	100).value
	Avail529	= sheet.cell(r,	101).value
	AvailRet	= sheet.cell(r,	102).value
	AvailIns	= sheet.cell(r,	103).value
	TaxMgd	= sheet.cell(r,	104).value
	Contrarn	= sheet.cell(r,	105).value
	BrkAvail	= sheet.cell(r,	106).value
	EnhIndex	= sheet.cell(r,	107).value
	Indexed	= sheet.cell(r,	108).value
	FoF	= sheet.cell(r,	109).value
	SocCons	= sheet.cell(r,	110).value
	NonDiv	= sheet.cell(r,	111).value
	ClsdAll	= sheet.cell(r,	112).value
	ClsdALLD	= sheet.cell(r,	113).value
	ClsdNew	= sheet.cell(r,	114).value
	ClsdNewD	= sheet.cell(r,	115).value
	FeeAccnt	= sheet.cell(r,	116).value
	FeeAdmin	= sheet.cell(r,	117).value
	FeeAdv	= sheet.cell(r,	118).value
	FeeAudit	= sheet.cell(r,	119).value
	FeeCust	= sheet.cell(r,	120).value
	FeeDistr	= sheet.cell(r,	121).value
	FeeIns	= sheet.cell(r,	122).value
	FeeLegal	= sheet.cell(r,	123).value
	FeeOrg	= sheet.cell(r,	124).value
	FeeOthr	= sheet.cell(r,	125).value
	FeePerf	= sheet.cell(r,	126).value
	FeeProf	= sheet.cell(r,	127).value
	FeeReg	= sheet.cell(r,	128).value
	FeeShRpt	= sheet.cell(r,	129).value
	FeeTA	= sheet.cell(r,	130).value
	Domicile	= sheet.cell(r,	131).value
	BaseCurr	= sheet.cell(r,	132).value
	ObsoltD	= sheet.cell(r,	133).value
	ObsoltTy	= sheet.cell(r,	134).value
	DiscCurr	= sheet.cell(r,	135).value
	ShrType	= sheet.cell(r,	136).value
	AABdLR	= sheet.cell(r,	137).value
	AABdL	= sheet.cell(r,	138).value
	AACshLR	= sheet.cell(r,	139).value
	AACBLR	= sheet.cell(r,	140).value
	AAEqLR	= sheet.cell(r,	141).value
	AAxUSBLR	= sheet.cell(r,	142).value
	AAxUSELR	= sheet.cell(r,	143).value
	AAOtLR	= sheet.cell(r,	144).value
	AAPSLR	= sheet.cell(r,	145).value
	AAUSBLR	= sheet.cell(r,	146).value
	AAUSELR	= sheet.cell(r,	147).value

	values = (Name, CUSIP, SecId, FundId, MRFundID, PerfID, Ticker, OldestSh, Advisor, FirmName, BrndName, BrndNmID, InceptDt, GlBrdGrp, GlBrdCat, USCatGrp, MSCat, MSCatID, MSCatStr, MSInsCat, MSRatAll, MSRat3Y, MSRat5Y, MSRat10Y, EqStyleL, EqStyleS, FxStyleL, MSSubAdv, MSSubFee, TeamMgd, MgrHist, MgrName, MgrTnrAv, MgrTnrLg, PPBnch, PPBnchID, PPBIncpt, SPBench, AUMShr, AUMShrDt, FSizeCur, FSizeDt, hldBondL, hldBondS, hldLong, hldOthrL, hldOthrS, hldStckL, hldStckS, hldT10, hldofInv, hldAT10, hldAT10D, Yld12M, Yld12MD, SECYld, SECYldD, DivdCur, DivdCurD, DivdNAVC, PortDate, VirtClss, FeeMgt, NExpSAR, NExpSARD, ChrgAR, ChrgARD, NExpPrs, NExpPrsD, ObjPros, Fee12b1, TurnRat, TurnRatD, CashN, AABondN, AACashN, AAEqN, AANUSBdN, AANUSEqN, AAOthrN, AAUSBdN, AAUSEqN, AAABN, AABondL, AACashL, AAEqL, AANUSBdL, AANUSEqL, AAOthrL, AAUSBdL, AAUSEqL, FoFIH, Leveragd, FeederFd, MasterFd, MasterN, MSTRFundID, Sharia, Ethical, ModelFd, Avail529, AvailRet, AvailIns, TaxMgd, Contrarn, BrkAvail, EnhIndex, Indexed, FoF, SocCons, NonDiv, ClsdAll, ClsdALLD, ClsdNew, ClsdNewD, FeeAccnt, FeeAdmin, FeeAdv, FeeAudit, FeeCust, FeeDistr, FeeIns, FeeLegal, FeeOrg, FeeOthr, FeePerf, FeeProf, FeeReg, FeeShRpt, FeeTA, Domicile, BaseCurr, ObsoltD, ObsoltTy, DiscCurr, ShrType, AABdLR, AABdL, AACshLR, AACBLR, AAEqLR, AAxUSBLR, AAxUSELR, AAOtLR, AAPSLR, AAUSBLR, AAUSELR)

	# print("inserting row ", r)
	mycursor.execute(sqlFormula, values)

print("finished, close db cursor.")
mycursor.close()

print("will commit to db...")
mydb.commit()
mydb.close()
print("success commit, close db.")


